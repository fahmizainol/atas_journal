"""Parse ATAS xlsx exports (Statistics / Journal / Executions) into DB rows.

Timezone handling: Journal and Executions timestamps are naive — they carry
the clock of whatever timezone ATAS was set to when the file was exported.
The importer tags those naives with ``source_tz`` (default America/New_York;
older exports used Asia/Kuala_Lumpur — override per-import) and stores both
the source-local ISO and the UTC ISO. The Statistics sheet is already UTC and
is stored verbatim.
"""

from __future__ import annotations

import hashlib
import sqlite3
from contextlib import AbstractContextManager, nullcontext
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

from . import db
from .config import (
    ET_TZ,
    IMPORTS_DIR,
    SOURCE_TZ_AFTER_SWITCH,
    SOURCE_TZ_BEFORE_SWITCH,
    SOURCE_TZ_SWITCH_DATE,
    UTC_TZ,
    normalize_instrument,
)

DEFAULT_SOURCE_TZ = ET_TZ


def _utc_iso(dt: datetime | None, source_tz: ZoneInfo) -> str | None:
    if dt is None:
        return None
    return dt.replace(tzinfo=source_tz).astimezone(UTC_TZ).isoformat()


def _local_iso(dt: datetime | None, source_tz: ZoneInfo) -> str | None:
    if dt is None:
        return None
    return dt.replace(tzinfo=source_tz).isoformat()


def _journal_key(row: dict) -> str:
    parts = [
        str(row["account"]), str(row["instrument"]),
        str(row["open_ts_local"]), str(row["close_ts_local"]),
        str(row["open_price"]), str(row["close_price"]), str(row["pnl"]),
    ]
    return hashlib.sha1("|".join(parts).encode()).hexdigest()


def source_key(path: Path, base: Path = IMPORTS_DIR) -> str:
    """The identity a file imports under: its path relative to the imports dir.

    Root files keep their bare name (so every pre-existing DB row is already
    keyed correctly); files under ``backtest/<folder>/`` become
    ``backtest/<folder>/<name>``, which makes the folders independent
    namespaces — ATAS reuses date-range filenames, so a backtest export must
    never collide with a same-named live/replay export in the root. Files
    outside ``base`` (e.g. one-off CLI imports) keep their bare name too.
    """
    try:
        return path.resolve().relative_to(base.resolve()).as_posix()
    except ValueError:
        return path.name


def _sheet_rows(wb, name: str) -> list[tuple]:
    if name not in wb.sheetnames:
        return []
    return list(wb[name].iter_rows(values_only=True))


def parse_file(
    path: Path, source_tz: ZoneInfo = DEFAULT_SOURCE_TZ
) -> dict[str, list[dict]]:
    """Return normalized {executions, journal, statistics} record lists.

    ``source_tz`` is the timezone the naive Journal/Executions timestamps were
    recorded in (i.e. whatever ATAS was set to at export time).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    source = source_key(path)

    executions: list[dict] = []
    for r in _sheet_rows(wb, "Executions")[1:]:
        if r[0] is None or r[3] is None:
            continue
        account, instrument, ts, exch_id, direction, price, volume, _route, comm = r[:9]
        executions.append({
            "exchange_id": str(exch_id),
            "account": str(account),
            "instrument": normalize_instrument(str(instrument)),
            "ts_local": _local_iso(ts, source_tz),
            "ts_utc": _utc_iso(ts, source_tz),
            "direction": str(direction),
            "price": float(price),
            "volume": float(volume),
            "commission": float(comm or 0),
            "source_file": source,
        })

    journal: list[dict] = []
    for r in _sheet_rows(wb, "Journal")[1:]:
        if r[0] is None:
            continue
        (account, instrument, open_t, open_p, open_v, close_t, close_p,
         close_v, price_pnl, profit_ticks, pnl, comment) = r[:12]
        if not close_t:
            # Position still open at export time — ATAS leaves Close time as
            # an empty string. Skip it: the closed trade arrives (with a real
            # close and a different dedupe key) in a later export.
            continue
        rec = {
            "account": str(account),
            "instrument": normalize_instrument(str(instrument)),
            "open_ts_local": _local_iso(open_t, source_tz),
            "close_ts_local": _local_iso(close_t, source_tz),
            "open_ts_utc": _utc_iso(open_t, source_tz),
            "close_ts_utc": _utc_iso(close_t, source_tz),
            "open_price": float(open_p),
            "open_volume": float(open_v),
            "close_price": float(close_p),
            "close_volume": float(close_v),
            "price_pnl": float(price_pnl) if price_pnl is not None else None,
            "profit_ticks": float(profit_ticks) if profit_ticks is not None else None,
            "pnl": float(pnl) if pnl is not None else None,
            "comment": str(comment or ""),
            "source_file": source,
        }
        rec["dedupe_key"] = _journal_key(rec)
        journal.append(rec)

    statistics: list[dict] = []
    stat_rows = _sheet_rows(wb, "Statistics")
    for r in stat_rows[1:]:
        if r[0] is None:
            continue
        metric = str(r[0])
        for scope, idx in (("Total", 1), ("Long", 2), ("Short", 3)):
            if idx < len(r) and r[idx] is not None:
                statistics.append({
                    "source_file": source,
                    "metric": metric,
                    "scope": scope,
                    "value": str(r[idx]),
                })

    return {"executions": executions, "journal": journal, "statistics": statistics}


def _disk_mtime_iso(path: Path) -> str:
    """The file's on-disk modified time as UTC ISO (Windows "Date modified")."""
    return datetime.fromtimestamp(path.stat().st_mtime, tz=UTC_TZ).isoformat()


def auto_source_tz_for_date(mdate: date) -> ZoneInfo:
    """Pick the source tz for an export modified on ``mdate``.

    Exports modified before ``SOURCE_TZ_SWITCH_DATE`` were taken while ATAS was
    set to Kuala Lumpur; from that date on, New York. See config for the switch.
    """
    return (
        SOURCE_TZ_AFTER_SWITCH
        if mdate >= SOURCE_TZ_SWITCH_DATE
        else SOURCE_TZ_BEFORE_SWITCH
    )


def _auto_source_tz(path: Path) -> ZoneInfo:
    """Auto-pick the source tz from a file's local "Date modified"."""
    return auto_source_tz_for_date(datetime.fromtimestamp(path.stat().st_mtime).date())


def store_parsed(
    conn: sqlite3.Connection,
    source_name: str,
    parsed: dict[str, list[dict]],
    file_mtime: str | None = None,
    mode: str | None = None,
    model_id: int | None = None,
) -> dict[str, int]:
    """Write one already-parsed export to the DB (the caller holds any lock)."""
    counts = {
        "executions": db.insert_executions(conn, parsed["executions"]),
        "journal": db.insert_journal(conn, parsed["journal"]),
        "statistics": db.insert_statistics(conn, parsed["statistics"]),
    }
    db.mark_imported(conn, source_name, file_mtime=file_mtime)
    # INSERT OR IGNORE: re-importing an export must not undo a mode set to
    # backtest, or un-archive a session the user archived.
    inferred, account = db.infer_session(r["account"] for r in parsed["journal"])
    db.upsert_session(conn, source_name, mode or inferred, account, model_id=model_id)
    return counts


def import_file(
    conn: sqlite3.Connection,
    path: Path,
    source_tz: ZoneInfo = DEFAULT_SOURCE_TZ,
    file_mtime: str | None = None,
    mode: str | None = None,
    model_id: int | None = None,
    lock: AbstractContextManager | None = None,
) -> dict[str, int]:
    """``mode``/``model_id`` override the inferred session mode — the watcher
    passes ``mode='backtest'`` plus the model its drop-box folder declares.

    ``lock`` guards only the DB writes: the slow openpyxl parse runs outside it
    so concurrent API requests aren't stalled behind an import pass.
    """
    parsed = parse_file(path, source_tz=source_tz)
    with lock if lock is not None else nullcontext():
        return store_parsed(
            conn, source_key(path), parsed,
            file_mtime=file_mtime, mode=mode, model_id=model_id,
        )


def import_dir(
    conn: sqlite3.Connection,
    directory: Path = IMPORTS_DIR,
    source_tz: ZoneInfo | None = None,
    lock: AbstractContextManager | None = None,
) -> dict[str, dict]:
    """Import every .xlsx in ``directory``.

    ``source_tz`` forces a single tz for every file. When ``None`` (the
    default), each file's tz is chosen from its modified date via
    :func:`_auto_source_tz` — KL before the switch, NY from it on.
    ``lock``, when given, is held per-file around the DB writes only.
    """
    results: dict[str, dict] = {}
    for path in sorted(Path(directory).glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        # Watched-dir files keep their real mtime, so read it straight off disk.
        tz = source_tz or _auto_source_tz(path)
        results[path.name] = import_file(
            conn, path, source_tz=tz, file_mtime=_disk_mtime_iso(path), lock=lock
        )
    return results
